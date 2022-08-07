# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import os
import datetime
from datetime import timedelta
from pytz import timezone, utc
import openpyxl
import requests

application=Flask(__name__)

# 참고 사항
# 변수명 앞에 d가 붙은 것은 저장된 데이터에서 불러온 값, d가 붙지 않은 것은 현재 or 입력한 데이터 값

KST=timezone('Asia/Seoul')
Days = ["일요일","월요일","화요일","수요일","목요일","금요일","토요일"] # 요일 이름
CompactDays = ['월','화','수','목','금','토','일'] #학사일정용
mealname = ["아침","점심","저녁"] # 식사 이름
mday = [31,28,31,30,31,30,31,31,30,31,30,31] # 매월 일 수
Msg = [["[오늘 아침]","[오늘 점심]","[오늘 저녁]"],["[내일 아침]","[내일 점심]","[내일 저녁]"]] # 급식 title
Menu = [["","",""],["","",""]] # 오늘, 내일 급식
Menu_saved_date = "" # 급식 불러온 날짜
GEvent = [{},{},{}]
Event_Saved_date = ""
mixed_table = []
timetable_save_date = ''
classn = ["11","12","13","14","21","22","23","24","31","32","33","34"] # 반 이름
classN = [20,20,20,21,20,20,19,19,13,13,13,13] # 반 학생 수
Name = [["강규준","김민석","김민재","김상준","김승민","김신정","김영훈","김용환","김준","김준희","남도균","박혜진","백정우","변승우","서상희","이경준","이예지","이은성","임혜빈","장덕명"],
        ["권오제","김동인","김민상","김예진","김예환","노영완","류지나","박정후","박준혁","박창현","서수환","엄강희","오민건","오수연","이민형","이성제","이수지","이승은","이형석","한재민"],
        ["김두현","김민규","김태윤","김희석","문도현","백나현","백준우","송하진","심규호","여지은","이민형","이서윤","이승재","이형민","임유택","장민석","정영훈","최시영","최예원","홍준서"],
        ["김기현","김민찬","김진송","도현송","박상민","박준원","배성훈","배채원","서지욱","송예송","양상현","오세현","이성혁","이윤서","이재경","임승윤","장정현","정민준","정지원","조예찬","조은솔"],
        ["곽성은", "구현우", "김도연", "김라영", "남민주", "박서우", "박현상", "배원호", "서준희", "서현", "염민호", "오은찬", "이어진", "임경규", "장혁진", "전성빈", "정민규", "조현준", "하도헌", "현준하"],
        ["김근형", "김서한", "김수민", "김정빈", "류현서", "박수현", "박종휘", "배성렬", "배창빈", "안도윤", "이주희", "이철욱", "장준영", "전재욱", "정재환", "조문경", "조현성", "채정현", "최경호", "최요훈"],
        ["곽민철", "김민석", "김수현", "김준우", "김태윤", "김하진", "박서진", "배준형", "배항준", "윤효상", "이근영", "이유진", "이재덕", "정혜인", "정호원", "조재현", "주선우", "편예준", "하승민"],
        ["강지운", "고정우", "권나리", "김건우", "김건호", "김민성", "김하준", "박지훈", "박한상", "서정현", "신수원", "양희정", "이지훈", "이효준", "정승현", "정태경", "차승빈", "허진호", "황우성"],
        ["권민철","김규리","김동현","김민준","김성윤","김승진","김유진","김진서","서원준","은성민","이동현","이영해","전민경"],
        ["고영건","김동은","노동완","민수현","박신후","박재영","박형준","백길홍","이동건","전서희","전제빈","정유라","정은주"],
        ["강호연","김민경","김민지","김세현","문재영","박재범","박주용","신은규","은태호","이준엽","장재혁","전우주","최승빈"],
        ["권인구","김동민","김재용","김태윤","박성민","박현제","용유성","유혜원","이민주","이지언","이효욱","주윤찬","최동하"]]
temp_timetables = {'월':[['','','',''],['','','',''],['','','','']],'화':[['','','',''],['','','',''],['','','','']],'수':[['','','',''],['','','',''],['','','','']],'목':[['','','',''],['','','',''],['','','','']],'금':[['','','',''],['','','',''],['','','','']]}
Teachers = {}
Selected_Subject = {}
Moveseat_saved_date = ""

def Make_aDay(L): #급식, 학사일정용 날짜 출력 함수
    now = datetime.datetime.now()
    after = now + timedelta(days = L)
    snow = now.strftime('%Y%m%d')
    safter = after.strftime('%Y%m%d')
    return snow,safter

def Make_aTlist():
    fr=open("/home/ubuntu/DG1Sbot2/teacher data.txt","r",encoding='UTF8') # 학번 불러오기
    lines=fr.readlines()
    global Teachers
    for line in lines:
        st = line.split(' ')
        Teachers[st[0]] = st[1].replace('\n','')
    fr.close()

#/home/ubuntu/DG1Sbot2/teacher data.txt

def Make_aSellist():
    fr=open("/home/ubuntu/DG1Sbot2/subject select.txt","r",encoding='UTF8') # 학번 불러오기
    lines=fr.readlines()
    global Selected_Subject
    for line in lines:
        sel = line.split(' ')
        Selected_Subject[sel[0]] = [sel[1],sel[2].replace('\n','')]
    fr.close()

#/home/ubuntu/DG1Sbot2/subject select.txt

def Temp_timetable():
    fr=open("/home/ubuntu/DG1Sbot2/time table data.txt","r",encoding='UTF8') # 학번 불러오기
    global temp_timetables
    lines=fr.readlines()
    Make_aTlist()
    Make_aSellist()
    res = []
    for line in lines:
        ttt = line.split('\t') #temp time table
        day = CompactDays[int(ttt[2]) - 1]
        grade = int(ttt[0]) - 1
        classN = int(ttt[1]) - 1 
        time = int(ttt[3]) - 1 
        subject = ttt[4].replace('\n','')
        try:
            res.append({'day':day,'grade':grade,'class':classN,'time':time,'subject':subject,'teacher':Teachers[str(grade+1)+subject]})
        except:
            res.append({'day':day,'grade':grade,'class':classN,'time':time,'subject':subject})
    fr.close()
    for items in res:
        if 'teacher' in list(items.keys()):
            if temp_timetables[items['day']][items['grade']][items['class']] == '':
                temp_timetables[items['day']][items['grade']][items['class']] = str(items['time'] + 1) + '교시  ' + items['subject'] + '  ' + items['teacher'] + 't' +'\n'
            else:
                temp = temp_timetables[items['day']][items['grade']][items['class']]
                temp = temp + str(items['time'] + 1) + '교시  ' + items['subject'] + '  ' + items['teacher'] + 't' +'\n'
                temp_timetables[items['day']][items['grade']][items['class']] = temp
        else:
            if temp_timetables[items['day']][items['grade']][items['class']] == '':
                temp_timetables[items['day']][items['grade']][items['class']] = str(items['time'] + 1) + '교시  ' + items['subject'] +'\n'
            else:
                temp = temp_timetables[items['day']][items['grade']][items['class']]
                temp = temp + str(items['time'] + 1) + '교시  ' + items['subject'] +'\n'
                temp_timetables[items['day']][items['grade']][items['class']] = temp


Temp_timetable()
#/home/ubuntu/DG1Sbot2/time table data.txt

def load_timetable(stid): # 시간표 출력 함수
    global temp_timetables,Selected_Subject
    now,after = Make_aDay(1)
    wd = datetime.datetime.strptime(now,'%Y%m%d')
    ugrade = int(stid[0]) -1
    uclassN = int(stid[1]) -1
    nowdays = wd.weekday()
    final_table = []
    final_date = [] #mixed_table == [] or
    if temp_timetables != {}:
        if nowdays == 0: dl = [0,1,2,3,4]
        if nowdays == 1: dl = [1,2,3,4,0]
        if nowdays == 2: dl = [2,3,4,0,1]
        if nowdays == 3: dl = [3,4,0,1,2]
        if nowdays == 4: dl = [4,0,1,2,3]
        if nowdays == 5 or nowdays ==6: dl = [0,1,2,3,4]
        for dd in dl:
            fday = list(temp_timetables.keys())[dd]
            final_date.append(fday)
            sub = temp_timetables[fday][ugrade][uclassN]
            if ugrade == 2:
                sel1 = Selected_Subject[stid][0]
                sel2 = Selected_Subject[stid][1]
                sub = sub.replace('선택1',sel1)
                sub = sub.replace('선택2',sel2)
            final_table.append(sub)
    final_TimeTable = [{"title": date , "description" : event} for date,event in zip(final_date,final_table)]
    
    return final_TimeTable

@application.route('/link', methods=['POST'])
def response_link(): # 온라인 시간표 대답 함수
        
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid="none"
    fr=open("/home/ubuntu/DG1Sbot2/user data.txt","r") # 학번 불러오기
    lines=fr.readlines()
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1];
        if dusid==userid: stid=dstid
    fr.close()

    if stid != 'none':
        items = load_timetable(stid)

    if stid=="none":
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "basicCard": {
                            "title": "[학번 등록]",
                            "description": "학번이 등록되어 있지 않습니다.\n아래 버튼을 눌러 학번을 등록해주세요",
                            "buttons": [ { "action": "message", "label": "학번 등록", "messageText": "학번 등록" } ]
                        }
                    }
                ]
            }
        }
    else :
        res={ # 답변
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "carousel": {
                                "type": "basicCard",
                                "items": items
                            }
                        },
                        {
                            "simpleText":{
                                "text": "* 실제 시간표는 위와 다를 수도 있다는 점 유의하시기 바랍니다."
                            }
                        }
                    ]
                }
            }
        return jsonify(res)

def what_is_menu():  # made by 1316, 1301 advanced by 2106
    global Menu, Menu_saved_date
    now, after = Make_aDay(1)
    if Menu_saved_date == "" or Menu_saved_date != now:
        Menu = [["", "", ""], ["", "", ""]]
        Menu_saved_date = now
        url = "https://open.neis.go.kr/hub/mealServiceDietInfo"
        params = {'KEY': 'b9558a909eb84bc68f5dd7add35f34a0',
                  'ATPT_OFCDC_SC_CODE': 'D10',
                  'SD_SCHUL_CODE': '7240331',
                  'MLSV_FROM_YMD': now,
                  'MLSV_TO_YMD': after,
                  'Type': 'json',
                  'pIndex': 1,
                  'pSize': 100}
        response = requests.get(url, params=params)
        res = response.json()
        try:
          for i in range(res['mealServiceDietInfo'][0]['head'][0]['list_total_count']):
              if res['mealServiceDietInfo'][1]['row'][i]['MLSV_YMD'] == now:
                  tempmenu = res['mealServiceDietInfo'][1]['row'][i]['DDISH_NM']
                  tempmenu = tempmenu.replace('(조)','') #불필요한 기호 제거
                  tempmenu = tempmenu.replace('(중)','')
                  tempmenu = tempmenu.replace('(석)','')
                  tempmenu = tempmenu.replace('#','')
                  tempmenu = tempmenu.replace('*','')
                  tempmenu = tempmenu.split('<br/>')
                  final_menu = "\n".join(tempmenu)
                  Menu[0][int(res['mealServiceDietInfo'][1]['row'][i]['MMEAL_SC_CODE']) - 1] = final_menu
              else:
                  tempmenu = res['mealServiceDietInfo'][1]['row'][i]['DDISH_NM']
                  tempmenu = tempmenu.replace('(조)','') #불필요한 기호 제거
                  tempmenu = tempmenu.replace('(중)','')
                  tempmenu = tempmenu.replace('(석)','')
                  tempmenu = tempmenu.replace('#','')
                  tempmenu = tempmenu.replace('*','')
                  tempmenu = tempmenu.split('<br/>')
                  final_menu = "\n".join(tempmenu)
                  Menu[1][int(res['mealServiceDietInfo'][1]['row'][i]['MMEAL_SC_CODE']) - 1] = final_menu
        except:
          Menu = [["", "", ""], ["", "", ""]]

    req = request.get_json()  # 파라미터 값 불러오기
    askmenu = req["action"]["detailParams"]["ask_menu"]["value"]

    now = datetime.datetime.utcnow()
    hour = int(utc.localize(now).astimezone(KST).strftime("%H"))  # Meal 계산
    minu = int(utc.localize(now).astimezone(KST).strftime("%M"))
    if (hour == 7 and minu >= 30) or (hour >= 8 and hour <= 12) or (hour == 13 and minu < 30):
        Meal = "아침"  # 가장 최근 식사가 언제인지 자동 계산
    elif (hour == 13 and minu >= 30) or (hour >= 14 and hour < 19) or (hour == 19 and minu < 30):
        Meal = "점심"
    else:
        Meal = "저녁"
    i = 0
    if Meal == "아침": fi = 1; si = 2; ti = 0  # 아침 점심 저녁 정보 불러오기 및 배열
    elif Meal == "점심": fi = 2; si = 0; ti = 1
    elif Meal == "저녁": fi = 0; si = 1; ti = 2
    if askmenu == "내일 급식": fi = 0; si = 1; ti = 2; i = 1
    first = Menu[i][fi]
    second = Menu[i][si]
    third = Menu[i][ti]
    if Menu[i][fi] == "": first = "등록된 급식이 없습니다."
    if Menu[i][si] == "": second = "등록된 급식이 없습니다."
    if Menu[i][ti] == "": third = "등록된 급식이 없습니다."
    return Msg[i][fi], Msg[i][si], Msg[i][ti], first, second, third


@application.route('/menu', methods=['POST'])
def response_menu():  # 메뉴 대답 함수
    msg1, msg2, msg3, menu1, menu2, menu3 = what_is_menu()
    if menu1 == "등록된 급식이 없습니다." and menu2 == "등록된 급식이 없습니다." and menu3 == "등록된 급식이 없습니다.":
        res = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": "급식이 없는 날입니다."
                        }
                    }
                ]
            }
        }
    else:
        res = {  # 답변
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "carousel": {
                            "type": "basicCard",
                            "items": [
                                {"title": msg1, "description": menu1},
                                {"title": msg2, "description": menu2},
                                {"title": msg3, "description": menu3}
                            ]
                        }
                    },
                    {
                      "simpleText": {
                      "text": ' (1.난류, 2.우유, 3.메밀, 4.땅콩, 5.대두, 6.밀, 7.고등어, 8.게, 9.새우, 10.돼지고기, 11.복숭아, 12.토마토, 13.아황산염, 14.호두, 15.닭고기, 16.쇠고기, 17.오징어, 18.조개류(굴,전복,홍합 등)'
                      }
                    }
                ]
            }
        }
    return jsonify(res)

@application.route('/colcheck', methods=['POST'])
def check_wp():
    
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid="none"
    fr=open("/home/ubuntu/DG1Sbot2/user data.txt","r") # 학번 불러오기
    lines=fr.readlines()
    fr.close()
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1];
        if dusid==userid: 
            stid=dstid
            break
    
    if stid=="none":
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "basicCard": {
                            "title": "[학번 등록]",
                            "description": "학번이 등록되어 있지 않습니다.\n아래 버튼을 눌러 학번을 등록해주세요",
                            "buttons": [ { "action": "message", "label": "학번 등록", "messageText": "학번 등록" } ]
                        }
                    }
                ]
            }
        }
    else :
        printmsg=""
        url = 'http://3.34.43.157:5000/colstdata'
        headers = { 'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.77 Safari/537.36'}
        response = requests.get(url,headers=headers) # url로부터 가져오기
        if response.status_code == 200: 
            source = response.text
            lines=source.split("\n")
            for line in lines:
                data=line.rstrip('\n').split(' ')
                if len(data)<4: continue
                datastid=data[0]
                datawarning=data[1]
                datapenalty=data[2]
                datareason=data[3:]
                if stid==datastid:
                    printmsg="[경고/벌점 현황]\n학번 : "+stid+"\n경고 "+datawarning+"회, 벌점 "+datapenalty+"점"
                    if len(datareason)!=1:
                        reasons=""
                        for reason in datareason:
                            if reason=="none": continue
                            reasons+="\n"+reason.replace('_',' ')[:10]+' '+reason.replace('_',' ')[10:]
                        printmsg+="\n사유 :"+reasons

        res={
            "version": "2.0",
            "template": {
                "outputs":[
                    {
                        "simpleText": {
                            "text": printmsg
                        }
                    }
                ]
            }
        }
    return jsonify(res)

def load_event(grade): #made by 2106
    global GEvent,Event_Saved_date
    now,after = Make_aDay(365)
    if Event_Saved_date == '' or GEvent == [] or Event_Saved_date != now:
        Event = []
        dictEvent = [[],[],[]] 
        GEvent = [{},{},{}]
        final_events = []
        Event_Saved_date = now
        url = "https://open.neis.go.kr/hub/SchoolSchedule"
        params = {'KEY' : 'b9558a909eb84bc68f5dd7add35f34a0',
                    'ATPT_OFCDC_SC_CODE':'D10',
                    'SD_SCHUL_CODE' : '7240331',
                    'AA_FROM_YMD':now,
                    'AA_TO_YMD':after,
                    'Type': 'json',
                    'pIndex': 1,
                    'pSize':200}
        response = requests.get(url,params = params)
        res = response.json() #일정을 가져옴
        for i in range(res['SchoolSchedule'][0]['head'][0]['list_total_count']): #{일정:날짜}
            if res['SchoolSchedule'][1]['row'][i]['EVENT_NM'] != '토요휴업일':
                G1,G2,G3 = False,False,False
                if res['SchoolSchedule'][1]['row'][i]['ONE_GRADE_EVENT_YN'] == 'Y': G1 = True
                if res['SchoolSchedule'][1]['row'][i]['TW_GRADE_EVENT_YN'] == 'Y': G2 = True
                if res['SchoolSchedule'][1]['row'][i]['THREE_GRADE_EVENT_YN'] == 'Y': G3 =True
                Event.append({'date':res['SchoolSchedule'][1]['row'][i]['AA_YMD'],'event':res['SchoolSchedule'][1]['row'][i]['EVENT_NM'],'grade':[G1,G2,G3]})
        for eve in Event:
            for i in range(3):
                if eve['grade'][i]:
                    dictEvent[i].append({eve['event']:eve['date']})
        for i in range(3):
            for j in range(len(dictEvent[i])):
                for event,date in dictEvent[i][j].items():
                    month = date[4:6]
                    if month[0] == '0': month = month[-1] + '월'
                    else: month = month + '월'
                    change = datetime.datetime.strptime(date, "%Y%m%d")
                    date = change.strftime('%-m월 %-d일') + '(' + CompactDays[change.weekday()] + ')'
                    if month not in GEvent[i]:
                        GEvent[i][month] = []
                        GEvent[i][month].append({event:date})
                        dlen = len(date)
                    else:
                        check = 0
                        for eve in GEvent[i][month]:
                            if event in eve.keys():
                                check = 1
                                eve[event] = eve[event][0:dlen] + ' ~ ' + date
                                break
                        if check == 0:
                            GEvent[i][month].append({event:date})
                            dlen = len(date)
    
    rawEvent = []
    for v1 in GEvent[grade].values():
        temp = ''
        for v2 in v1:
            for k,v3 in v2.items():
              temp = temp + v3 + '  ' + k + '\n'
        rawEvent.append(temp)
    final_events = [{"title": month , "description" : event} for month,event in zip(GEvent[0].keys(),rawEvent)]       
    return final_events

@application.route('/eventcheck', methods=['POST']) #학사일정 대답 함수
def Evecheck():
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid="none"
    fr=open("/home/ubuntu/DG1Sbot2/user data.txt","r") # 학번 불러오기
    lines=fr.readlines()

    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1];
        if dusid==userid: stid=dstid
    fr.close()

    if stid != 'none':
        grade = int(stid[0]) -1
        final_events = load_event(grade)

    if stid=="none":
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "basicCard": {
                            "title": "[학번 등록]",
                            "description": "학번이 등록되어 있지 않습니다.\n아래 버튼을 눌러 학번을 등록해주세요",
                            "buttons": [ { "action": "message", "label": "학번 등록", "messageText": "학번 등록" } ]
                        }
                    }
                ]
            }
        }
    else:
        if final_events != []:
                res = {
                    "version": "2.0",
                    "template": {
                        "outputs": [
                            {
                                "carousel": {
                                    "type": "basicCard",
                                    "items": final_events
                                }
                            }
                        ]
                    }
                }
        else:
                res = {
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": "저장된 일정이 없습니다."
                        }
                    }
                ]
            }
        }
    return jsonify(res)

@application.route('/movepl', methods=['GET','POST'])
def Getuseranswerpl(): #유저 인풋 받기
    req = request.get_json()  # 파라미터 값 불러오기
    userid = req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    moving_place = req["action"]["detailParams"]["move_position"]["value"]
    moving_time = req["action"]["detailParams"]["move_time"]["value"]
    moving_together = req["action"]["detailParams"]["move_together"]["value"]
    move_reason = req["action"]["detailParams"]["move_reason"]["value"]
    stid = "none"

    now, after = Make_aDay(1)
    if Moveseat_saved_date == "" or Moveseat_saved_date != now:
        Moveseat_saved_date = now
        fr = open("final save.txt", "w", encoding='UTF8')  # 엑셀 채워 넣기
        fr.write("13\n")
        fr.close()

    fr = open("/home/ubuntu/DG1Sbot2/user data.txt", "r")  # 학번 불러오기
    lines = fr.readlines()
    for line in lines:
        datas = line.split(" ")
        dusid = datas[0]
        dstid = datas[1]
        if dusid == userid: stid = dstid
    fr.close()

    comp = datetime.time(16, 40, 00)
    now = datetime.datetime.now().time()
    if stid=="none":
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "basicCard": {
                            "title": "[학번 등록]",
                            "description": "학번이 등록되어 있지 않습니다.\n아래 버튼을 눌러 학번을 등록해주세요",
                            "buttons": [ { "action": "message", "label": "학번 등록", "messageText": "학번 등록" } ]
                        }
                    }
                ]
            }
        }

    else:
        if comp < now:
            res = {
                "version": "2.0",
                "template": {
                    "outputs": [{"simpleText": {"text": "자리이동 가능 시간이 아닙니다\n 16시 40분 전까지 설문을 완료해 주세요"}}]
                }
            }
        else:
            res = {
                "version": "2.0",
                "template": {
                    "outputs": [
                        {
                            "basicCard": {
                                "title": "[저장 완료]",
                                "description": "이동 시간: " + moving_time + "\n" + "이동 장소: " + moving_place + "\n"+ "이동 인원: "+ stid+ " "+ moving_together +"\n"+"이동 사유: " + move_reason
                            }
                        }
                    ]
                }
            }
            fw = open("/home/ubuntu/DG1Sbot2/final save.txt", "a")
            fw.write(moving_place + "\t" + moving_time + "\t" + stid + " " +  moving_together + "\t" + move_reason + "\n")
            fw.close()
    return jsonify(res)

def to_excel():
    wb = openpyxl.load_workbook('자리이동 명렬.xlsx', data_only=True)  # 엑셀 기본 형식
    fr = open("final save.txt", "r", encoding='UTF8')  # 엑셀 채워 넣기
    for pg in range(3):                         #엑셀 파일 초기화
        for col in range(4):
            if pg == 0:
                i = col + 0
            elif pg == 1:
                i = col + 4
            elif pg == 2:
                i = col + 8
            for row in range(3,len(Name[i])):
                sheet = wb[str(pg+1) + "학년"]
                sheet.cell(row, col).value = None
    lines = fr.readlines()
    for line in lines:
        if line == lines[0]: continue
        if "none" in line: continue
        datas = line.split("\t")
        peo = datas[2]
        per = peo.split(" ")
        mvPL = datas[0]
        mvTi = datas[1]
        mvRe = datas[3].replace("\n", "")
        for p in per:
            if p != '.':
                sn = p[0] + "학년"
                row = int(p[2:]) + 2
                col = int(p[1]) * 2 + 1
                sheet = wb[sn]
                if sheet.cell(row, col).value == None:
                    sheet.cell(row, col).value = mvTi + "    " + mvPL + "    " + mvRe
                else:
                    sv = sheet.cell(row, col).value.split("    ")
                    if sv[0] != mvTi:
                        sheet.cell(row, col).value = "    ".join(
                            sv[len(sv) - 3:]) + "    " + mvTi + "    " + mvPL + "    " + mvRe
                    else:
                        del sv[len(sv) - 3:]
                        sheet.cell(row, col).value = "    ".join(
                            sv[len(sv) - 3:]) + "    " + mvTi + "    " + mvPL + "    " + mvRe

    fr.close()
    return None


@application.route('/stid', methods=['POST'])
def input_stid(): # 학번 입력 함수
        
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid=req["action"]["detailParams"]["student_id"]["value"]
    check=False
    
    fr=open("/home/ubuntu/DG1Sbot2/user data.txt","r") # userdata 저장 및 변경
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/DG1Sbot2/user data.txt","w")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]
        if dusid==userid: 
            fw.write(userid+" "+stid+" 7 none 0 none none\n")
            check=True
        else : fw.write(line)
    if check==False : fw.write(userid+" "+stid+" 7 none 0 none none\n")
    fw.close()
    res={
        "version": "2.0",
        "template": { "outputs": [ { "simpleText": { "text": "학번이 "+stid+"(으)로 등록되었습니다." } } ] }
    }
    return jsonify(res)
 
@application.route('/')
def index():
    return render_template("index.html")

filename=""

@application.route('/texteditor')
def text_editor(): # 원하는 파일 사이트에서 보여주고 편집
    global filename
    filename=request.args.get('filename')
    fr=open("/home/ubuntu/DG1Sbot2/"+filename+".txt","r")
    data_send=fr.readlines()
    fr.close()
    if filename=="user data": data_send.sort(key=lambda x:x[13:17]) # 학번 순 정렬
    return render_template("texteditor.html",data=data_send, name=filename)

@application.route('/filesave', methods=['GET','POST'])
def save_as_file(): # txt file 저장하기
    if request.method=='POST':
        fr=open("/home/ubuntu/DG1Sbot2/"+filename+".txt","r")
        before=fr.read()
        fr.close()
        
        text=request.form['content']
        text=str(text)
        with open("/home/ubuntu/DG1Sbot2/"+filename+".txt","w",encoding='utf-8') as f:
            f.write(text)
    
        now=datetime.datetime.utcnow()
        hour=utc.localize(now).astimezone(KST).strftime("%H")
        minu=utc.localize(now).astimezone(KST).strftime("%M")
        date=utc.localize(now).astimezone(KST).strftime("%d")
        month=utc.localize(now).astimezone(KST).strftime("%m")
        year=utc.localize(now).astimezone(KST).strftime("%Y")
        fw=open("/home/ubuntu/DG1Sbot2/log.txt","a")
        fw.write('['+year+'-'+month+'-'+date+' '+hour+':'+minu+"] '"+filename+".txt' saved (Below is the contents before saving.)\n")
        fw.write(before+'\n')
        fw.close()
        
        return render_template("saved.html")
  
@application.route('/xlsave', methods=['GET','POST'])
def save_as_xlfile(): # file 저장하기
    if request.method == 'POST':
        f=request.files['xlfile']
        f.save("/home/ubuntu/DG1Sbot2/"+secure_filename(f.filename))
        return render_template("saved.html")

@application.route('/xldownload',  methods=["GET", "POST")
def download():
    return send_file("/home/ubuntu/DG1Sbot2/자리이동 명렬.xlsx", as_attachment=True)
  
@application.route('/dnldfile', methods=['GET','POST'])
def download_file(): # file 다운받기
    if request.method == 'POST':
        filename=request.form['downloadfilename']
        return send_file("/home/ubuntu/DG1Sbot2/"+filename, attachment_filename=filename, as_attachment=True)

@application.route('/file')
def upload_n_download():
    files=os.listdir("/home/ubuntu/DG1Sbot2")
    folders=[]
    for file in files:
        if not '.' in file: folders.append(file)
    for folder in folders:
        files.remove(folder)
    return render_template("file.html", files=files)

@application.route('/status')
def record_status():
    index=int(request.args.get('index'))
    n=classN[index]

    stid=[]
    for i in range(1,classN[index]+1):
        id=classn[index]
        if i<10: id+='0'
        id+=str(i)
        stid.append(id)
    
    name=Name[index]

    record=[]
    for i in range(25):
        record.append([])
        for j in range(13):
            record[i].append('')
        record[i].append(0)
    fr=open("/home/ubuntu/DG1Sbot2/final save.txt", "r")
    lines=fr.readlines()
    for line in lines:
        if line==lines[0]: continue
        if "none" in line: continue
        datas=line.split(' '); id=datas[0]; day=int(datas[1]); meal=int(datas[2]); seat=datas[3]
        if id[:2]==classn[index]:
            if 3*day+meal-4<0 or 3*day+meal-4>12: continue
            if record[int(id[2:4])-1][3*day+meal-4]=='': record[int(id[2:4])-1][13]+=1
            if seat==".": seat="X"
            record[int(id[2:4])-1][3*day+meal-4]=seat
    fr.close()
    mealN=int(lines[0].rstrip('\n'))
    for i in range(n):
        record[i][13]=str(round((record[i][13]/mealN)*100))+'%'
    
    return render_template("status.html", n=n, stid=stid, name=name, record=record)


@application.route('/ball')
def ball():
    return render_template("Ball.html")

if __name__ == "__main__":
    application.run(host='0.0.0.0', port=5000)
